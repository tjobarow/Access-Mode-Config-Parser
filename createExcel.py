import openpyxl, datetime
from openpyxl.styles import Font, Alignment, Border, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.workbook import workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import os, glob, json


def createExcel():
    #Create new workbook
    workbook = Workbook()

    #Define the Current Path and the Sub-directory containing configs
    cur_path = os.path.dirname(__file__)+"/customer_configs/"
    #For each JSON file in the directory
    for json_file_path in glob.glob(os.path.join(cur_path,'*.json')):
        
        #Split path on "/" so we can get hostname in next step
        path_split = json_file_path.split("/")
        #Then we take the last index of the split path (will be filename)
        # and split it on ".". We take the first index from that, which will be our hostname. 
        hostname = path_split[len(path_split)-1].split(".")[0]

        #Open JSON file
        json_file_obj = open(json_file_path,'r')
        #Load the file object as a dictionary
        json_file = json.load(json_file_obj)
        #print(json.dumps(json_file,indent=4))

        # Create a sheet based on hostname
        ws = workbook.create_sheet(hostname)

        #Since worksheets start as 1x1, we need to add
        # Two columns so we have 3 in total
        ws.insert_cols(0,2)
        # rows equal to the length of the dictionary, the pre-existing 1st row will be headers
        ws.insert_rows(0,(len(json_file)))

        #Set headers
        ws['A1'].value = "Interface"
        ws['B1'].value = "Current State"
        ws['C1'].value = "Desired State"

        #Set width
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 50

        #Openpyxl has rows start from 1, not 0. 
        # We start at 2 since 1 has the headers. 
        # We need this variable to increment to the next row
        row_index = 2
        #For each interface object in the json file. (Key contains dict key, while value holds dict
        # that contains the info we need)
        for key,value in json_file.items():
            print(hostname+ " - " + str(value))
            #Set cell A{row_index} (ex: A2 if we are on 2nd row) to have the interface name
            # We split on a whitespace character because the value in the dict is 
            # "interface FastEthernet4/0/34" for example, but we just want 
            # FastEthernet4/0/34, removing the interface keyword. 
            ws[f'A{str(row_index)}'].value = value['interface_name'].split(" ")[1]
            #If dot1x is already configured on the interface, we put its current mode in 
            # the current state column, and do the same for the desired state
            if value['dot1x_configured']:
                ws[f'B{str(row_index)}'].value = value['dot1x_mode']
                ws[f'C{row_index}'].value = value['dot1x_mode']
            # If dot1x is not configured, we put N/A as the interface does not have it configured.
            # We list the desired state as open, as we do not want to go directly into closed mode.
            else:
                ws[f'B{str(row_index)}'].value = "N/A"
                ws[f'C{str(row_index)}'].value = "open"
            #Move up a row. 
            row_index+=1

    if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])

            workbook_name = "Interface_Audit_Results.xlsx"

            print(f"~~~~~~~~~~~~~~~~~~~~~~~~~~~~ SAVING {workbook_name} ~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
            
            workbook.save(workbook_name)


if __name__ == "__main__":
    createExcel()